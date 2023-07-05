import openpyxl
class HomePageData:

    #test_HomePage_data = [{"name":"Shyam", "email": "test@test.com", "password":"password", "gender":"Male"},
      #                  {"name":"Gorasia", "email":"test2@test.com", "password":"password2", "gender":"Female"}]

# Integrate excelDemo into framework - Excel data driven

    @staticmethod
    def getTestData(testCaseName): #self only neeed if its not static method
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\shyam.gorasia\\PycharmProjects\\Book1.xlsx")
        sheet = book.active

        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == testCaseName:

                for j in range(2, sheet.max_column + 1):
                    # Dict["name"]="Gary Penn" - this same as below, below is better way of doing it
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i,
                                                                         column=j).value  # this will get the value in column Testcase1

        return[Dict] # this will return as a list - this is required when sending the dictionary
        #print(Dict)