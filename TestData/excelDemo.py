# How to drive data from Excel, instead of hardcoding into code.

import openpyxl

# initialises dictionary
Dict = {}

# this loads the excel workbook file
book = openpyxl.load_workbook("C:\\Users\\shyam.gorasia\\PycharmProjects\\Book1.xlsx")

# workbook has many sheets, so select an active sheet
sheet = book.active

# in python each row will be treated as starting from first row/column etc.

cell = sheet.cell(row=1, column=2)

# this extracts the value from that cell 1, 2
print(cell.value)
# another way of doing it
print(sheet['A5'].value)

# this will write back to the file
sheet.cell(row=2, column=2).value = "MyName"
print(sheet.cell(row=2, column = 2).value)

# how to know no. of rows and columns in sheet
print(sheet.max_row)
print(sheet.max_column)

# another way of doing it
print(sheet['A5'].value)

# For loop - how to print first row of values
for i in range(1, sheet.max_row+1): # this prints from 1 - 7, if you add +1
    print(sheet.cell(row=i, column=1).value)

    # this will start from 1,1 and increase row by 1

print("------ 2nd for loop -------")
# For loop - how to print all values
for i in range(1, sheet.max_row+1): #
    for j in range(1, sheet.max_column+1): # this will repeat 4 times as there are 4 columns before starting another row
        print(sheet.cell(row=i, column=j).value)

print("--------- printing certain test cases ---------")
# only printing certain test case values
for i in range(1, sheet.max_row+1): #
    if sheet.cell(row=i, column=1).value == "Testcase2": # to get rows - this will scan only if first value matches

        for j in range(1, sheet.max_column+1): # to get columns
            print(sheet.cell(row=i, column=j).value)

# we only need the data for each testcase, not "testcase1, 2 etc. So just change 1 to 2 in to get columns for loop

print("-----------------Loading the data from file into Dictionary--------------")
#------------------- Loading the data from file into Dictionary -------------


for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "Testcase1":

        for j in range(2, sheet.max_column+1):
            #Dict["name"]="Gary Penn" - this same as below, below is better way of doing it
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value #this will get the value in column Testcase1

print(Dict)